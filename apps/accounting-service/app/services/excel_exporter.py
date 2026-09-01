"""
Xuất báo cáo P&L ra file Excel — Mục 3.6.1 spec: "Xuất báo cáo Excel/PDF
(dùng openpyxl, pandas — Python mạnh hơn Node ở đây)".

Phase 4: lưu file tạm cục bộ rồi trả path để caller (Celery task hoặc API
endpoint) upload lên Object Storage (Cloudflare R2 — Mục 9.2 spec). Việc
upload thật cần thêm SDK S3-compatible — để TODO rõ ràng vì cần credentials
thật, không giả lập ở đây.
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from app.services.pnl_calculator import calculate_pnl

EXPORT_DIR = os.environ.get("EXPORT_TMP_DIR", "/tmp/369-reports")


def export_pnl_to_excel(store_id: str, period_start: datetime, period_end: datetime) -> str:
    os.makedirs(EXPORT_DIR, exist_ok=True)
    result = calculate_pnl(store_id, period_start, period_end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo Lãi Lỗ"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

    ws.append(["BÁO CÁO LÃI LỖ — 369 PLATFORM"])
    ws.append([f"Kỳ báo cáo: {period_start.date()} → {period_end.date()}"])
    ws.append([])
    ws.append(["Chỉ tiêu", "Giá trị (VNĐ)"])
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill

    ws.append(["Tổng doanh thu", result["total_revenue"]])
    ws.append(["Tổng chi phí", result["total_expense"]])
    ws.append(["Lợi nhuận ròng", result["net_profit"]])
    ws.append([])
    ws.append(["Chi tiết chi phí theo hạng mục"])
    for category, amount in result["expense_by_category"].items():
        ws.append([category, amount])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    filename = f"pnl_{store_id}_{period_start.date()}_{period_end.date()}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return filepath
